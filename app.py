#!/usr/bin/env python3
"""
Flask Web Application for Document Generation with Salesperson Management

Features:
  - Lists available DOCX templates (excluding any sanitized copies).
  - Displays a table of salespersons from an XLSX file (salespersons.xlsx).
  - After selecting a template, presents a form to fill in template placeholders
    (excluding Salesperson_Name, Salesperson_Email, Salesperson_Phone) and a dropdown
    to choose a salesperson.
  - Automatically injects the chosen salesperson's details into the document context.
  - Generates the final DOCX file with a filename based on "Client Company Name" and "Proposal date".
 
Dependencies:
  - Python 3.9+
  - Flask
  - docxtpl
  - openpyxl
  - (Optional) docx2pdf for PDF conversion
"""

import os
import re
import sys
import glob
import zipfile
import tempfile
import shutil
from flask import Flask, render_template, request, redirect, url_for, send_from_directory, flash
from docxtpl import DocxTemplate
# from docx2pdf import convert  # Uncomment if PDF conversion is desired and supported on your server
from openpyxl import Workbook, load_workbook

app = Flask(__name__)
app.secret_key = "supersecretkey"  # Required for flash messages

# Directories and files
BASE_DIR = os.path.abspath(os.path.dirname(__file__))
ORIGINAL_TEMPLATES_DIR = BASE_DIR  # Adjust if original templates are stored elsewhere
SANITIZED_DIR = os.path.join(BASE_DIR, "sanitized_templates")
OUTPUT_DIR = os.path.join(BASE_DIR, "output")
SALESPEOPLE_FILE = os.path.join(BASE_DIR, "salespersons.xlsx")

# Ensure required folders exist.
os.makedirs(SANITIZED_DIR, exist_ok=True)
os.makedirs(OUTPUT_DIR, exist_ok=True)

# --------------------
# Salesperson XLSX Utilities
# --------------------

def init_salespeople_file():
    """Create salespersons.xlsx with headers if it does not exist."""
    if not os.path.exists(SALESPEOPLE_FILE):
        wb = Workbook()
        ws = wb.active
        ws.append(["Name", "Email", "Phone"])
        wb.save(SALESPEOPLE_FILE)

def get_salespeople():
    """Return a list of salespersons as dictionaries from salespersons.xlsx."""
    init_salespeople_file()
    wb = load_workbook(SALESPEOPLE_FILE)
    ws = wb.active
    salespeople = []
    # Skip header row (first row)
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row and any(row):
            sp = {"Name": row[0] or "", "Email": row[1] or "", "Phone": row[2] or ""}
            salespeople.append(sp)
    return salespeople

# --------------------
# Document Generation Utilities
# --------------------

def extract_placeholders_from_xml(docx_path):
    """
    Extract placeholders from the document.xml inside the DOCX.
    Returns a list of unique placeholder strings.
    """
    try:
        with zipfile.ZipFile(docx_path, "r") as z:
            xml_content = z.read("word/document.xml").decode("utf-8")
    except Exception as e:
        print(f"Error reading the DOCX file: {e}")
        return []
    raw_matches = re.findall(r'{{(.*?)}}', xml_content, re.DOTALL)
    placeholders = set()
    for match in raw_matches:
        cleaned = re.sub(r'<[^>]+>', '', match).strip()
        if cleaned:
            placeholders.add(cleaned)
    return list(placeholders)

def sanitize_placeholder(placeholder):
    """
    Convert a placeholder to a valid Python identifier by replacing non-alphanumeric characters with underscores.
    """
    sanitized = re.sub(r'\W+', '_', placeholder)
    return sanitized.strip('_')

def sanitize_template_xml(template_path, mapping, sanitized_dir):
    """
    Create a sanitized version of the DOCX template.
    For each placeholder in the DOCX that, when cleaned, matches one of the keys in mapping,
    replace it with the corresponding sanitized version.
    The new DOCX is stored in sanitized_dir.
    Returns the path to the new sanitized template.
    """
    os.makedirs(sanitized_dir, exist_ok=True)
    temp_dir = tempfile.mkdtemp()
    with zipfile.ZipFile(template_path, 'r') as zin:
        zin.extractall(temp_dir)
    doc_xml_path = os.path.join(temp_dir, "word", "document.xml")
    try:
        with open(doc_xml_path, "r", encoding="utf-8") as f:
            xml_content = f.read()
    except Exception as e:
        print(f"Error reading document.xml: {e}")
        shutil.rmtree(temp_dir)
        return None
    def replacement(match):
        full_match = match.group(0)
        inner = match.group(1)
        cleaned = re.sub(r'<[^>]+>', '', inner).strip()
        if cleaned in mapping:
            return "{{" + mapping[cleaned] + "}}"
        else:
            return full_match
    new_xml = re.sub(r'{{(.*?)}}', replacement, xml_content, flags=re.DOTALL)
    try:
        with open(doc_xml_path, "w", encoding="utf-8") as f:
            f.write(new_xml)
    except Exception as e:
        print(f"Error writing modified document.xml: {e}")
        shutil.rmtree(temp_dir)
        return None
    sanitized_template_path = os.path.join(sanitized_dir, "sanitized_" + os.path.basename(template_path))
    with zipfile.ZipFile(sanitized_template_path, "w", zipfile.ZIP_DEFLATED) as zout:
        for foldername, subfolders, filenames in os.walk(temp_dir):
            for filename in filenames:
                file_path = os.path.join(foldername, filename)
                arcname = os.path.relpath(file_path, temp_dir)
                zout.write(file_path, arcname)
    shutil.rmtree(temp_dir)
    return sanitized_template_path

def get_value_case_insensitive(dictionary, target_key, default):
    """Retrieve a value from a dictionary using case-insensitive key matching."""
    for key, value in dictionary.items():
        if key.strip().lower() == target_key.strip().lower():
            return value.strip() if value.strip() else default
    return default

# --------------------
# Flask Routes
# --------------------

@app.route('/')
def index():
    """Home page: List available original templates and display the salesperson table."""
    # List only original templates (exclude any starting with "sanitized_")
    template_files = [f for f in glob.glob(os.path.join(ORIGINAL_TEMPLATES_DIR, "*Template*.docx"))
                      if not os.path.basename(f).startswith("sanitized_")]
    salespeople = get_salespeople()
    return render_template("index.html",
                           template_files=template_files,
                           salespeople=salespeople)

@app.route('/select_template', methods=["POST"])
def select_template():
    """
    After selecting a template, extract its placeholders and show a form to fill them.
    Exclude placeholders for Salesperson details (they will be auto-filled).
    Also include a dropdown to choose a salesperson.
    """
    template_file = request.form.get("template_file")
    if not template_file:
        flash("No template selected.")
        return redirect(url_for("index"))
    placeholders = extract_placeholders_from_xml(template_file)
    if not placeholders:
        flash("No placeholders found in the selected template.")
        return redirect(url_for("index"))
    placeholders.sort()
    # Filter out salesperson placeholders from the manual entry list.
    filtered_placeholders = [ph for ph in placeholders
                             if ph not in ["Salesperson_Name", "Salesperson_Email", "Salesperson_Phone"]]
    salespeople = get_salespeople()
    return render_template("fill_placeholders.html",
                           template_file=template_file,
                           placeholders=filtered_placeholders,
                           salespeople=salespeople)

@app.route('/generate_document', methods=["POST"])
def generate_document():
    """
    Process form data, generate the document, and provide a download link.
    The chosen salesperson’s details are automatically injected into the context.
    """
    template_file = request.form.get("template_file")
    if not template_file:
        flash("Template file missing.")
        return redirect(url_for("index"))
    # Re-extract the placeholders from the template (if needed)
    all_placeholders = extract_placeholders_from_xml(template_file)
    all_placeholders.sort()
    mapping = {}
    raw_values = {}
    context = {}
    # Process only non-salesperson placeholders from the form.
    for ph in all_placeholders:
        if ph in ["Salesperson_Name", "Salesperson_Email", "Salesperson_Phone"]:
            continue  # Skip these—will be auto-filled.
        key = sanitize_placeholder(ph)
        mapping[ph] = key
        value = request.form.get(key)
        raw_values[ph] = value
        context[key] = value
    # Get the selected salesperson from the dropdown.
    selected_salesperson = request.form.get("salesperson")
    if selected_salesperson:
        salespeople = get_salespeople()
        sp = next((s for s in salespeople if s["Name"].strip().lower() == selected_salesperson.strip().lower()), None)
        if sp:
            context["Salesperson_Name"] = sp["Name"]
            context["Salesperson_Email"] = sp["Email"]
            context["Salesperson_Phone"] = sp["Phone"]
    # Create a sanitized version of the template.
    sanitized_template_path = sanitize_template_xml(template_file, mapping, SANITIZED_DIR)
    if not sanitized_template_path:
        flash("Error sanitizing the template.")
        return redirect(url_for("index"))
    try:
        doc = DocxTemplate(sanitized_template_path)
        doc.render(context)
    except Exception as e:
        flash("Error rendering the document.")
        return redirect(url_for("index"))
    # Build output filename using raw values for "Client Company Name" and "Proposal date"
    client_name = get_value_case_insensitive(raw_values, "Client Company Name", "UnknownClient")
    proposal_date = get_value_case_insensitive(raw_values, "Proposal date", "UnknownDate")
    filename_base = f"Proposal_{client_name}_{proposal_date}"
    output_filename = filename_base + ".docx"
    output_path = os.path.join(OUTPUT_DIR, output_filename)
    try:
        doc.save(output_path)
    except Exception as e:
        flash("Error saving the generated document.")
        return redirect(url_for("index"))
    return render_template("result.html", output_filename=output_filename)

@app.route('/download/<filename>')
def download(filename):
    """Serve a generated file for download."""
    return send_from_directory(OUTPUT_DIR, filename, as_attachment=True)

# --------------------
# Run the App
# --------------------
if __name__ == '__main__':
    app.run(debug=True)
